"""The Worksheet class abstracts an Excel worksheet, capturing e.g. its header
row number, headers, data, etc"""

# --- Standard Library Imports ------------------------------------------------
import tkinter as tk
from pathlib import Path
from tkinter import filedialog
import collections

# --- Third Party Imports -----------------------------------------------------
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# --- Intra-Package Imports ---------------------------------------------------
from mentormatch import config
from mentormatch.import_worksheet import string_analysis


def get_path():
    root = tk.Tk()
    root.withdraw()
    path = Path(filedialog.askopenfilename())
    return path


def get_workbook(path):
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except (FileNotFoundError, InvalidFileException):
        raise config.MentormatchError(f"\nError: '{path} ' is not a valid excel path.")


def get_worksheet(workbook, worksheet_name):
    try:
        return workbook[worksheet_name]
    except KeyError:
        raise config.MentormatchError(f"\nError: Workbook does not contain a '{worksheet_name}' worksheet")


def get_header_row_number(worksheet, header_names, max_row=20):
    """Find the row that matches the header_names. Max: 20 rows"""
    rows = [get_worksheet_row(worksheet, row) for row in range(1, max_row + 1)]
    best_index = string_analysis.find_most_similar_string_sequence(header_names, rows)
    return best_index + 1


def get_worksheet_row(worksheet, row_number) -> list:
    return [worksheet.cell(row_number, col) for col in range(1, worksheet.max_col)]


def get_header_names(worksheet, header_row_number) -> list:
    """Return header names, ignoring blank cells. Raise error if there are duplicates."""
    header_names = [
        header for header in get_worksheet_row(worksheet, header_row_number)
        if header  # We ignore any blank cells
    ]
    header_counts = collections.Counter(header_names)
    repeated_headers = [header for header in header_names if header_counts[header] > 1]
    if repeated_headers:
        msg = f'\nError: the following headers in {worksheet.name} have duplicates:\n{repeated_headers}'
        raise config.MentormatchError(msg)
    if 'row' in header_names:
        # This is because 'row' will be a automatically generated value
        msg = f"\nError: 'row' is not a valid header in {worksheet.name}"
        raise config.MentormatchError(msg)
    return header_names


def convert_rows_to_dicts(worksheet, header_row_number, header_names=None):

    # --- get headers ---------------------------------------------------------
    actual_header_row = get_worksheet_row(worksheet, header_row_number)

    # --- check worksheet for desired headers ---------------------------------
    if header_names:
        missing_headers = set(header_names) - set(actual_header_row)
        if missing_headers:
            msg = f'\nError: the following headers in {worksheet.name} are missing:\n{missing_headers}'
            raise config.MentormatchError(msg)
    else:
        header_names = actual_header_row

    # --- select cols ---------------------------------------------------------
    cols = [actual_header_row.index(header) for header in header_names]

    # --- get dicts -----------------------------------------------------------
    records = list()
    for row in range(header_row_number + 1, worksheet.max_row):
        record = dict()
        record['row'] = row
        for header_index, col in enumerate(cols):
            header = header_names[header_index]
            value = worksheet.cell(row, col)
            record[header] = value
        records.append(record)
    return records



# class Worksheet:
#
#
#
#         if get_header_row_number and converters:
#             header_row = header.get_header_row_number(excel_path, excel_sheet_name, converters.keys())
#
#         # --- raise exceptions for missing file, worksheet, or headers --------
#         # --- print warnings for extra headers --------------------------------
#         try:
#             actual_cols = pd.read_excel(
#                 io=excel_path,
#                 sheet_name=excel_sheet_name,
#                 header=header_row-1,
#                 nrows=0,
#             ).columns
#         except FileNotFoundError:
#             raise config.MentormatchError(f'<{excel_path}> not valid file.')
#         except xlrd.biffh.XLRDError:
#             raise config.MentormatchError(f"<{excel_sheet_name}> sheet not found")
#         if converters is not None:
#             expected_cols = converters.keys()
#             missing_cols = [col for col in expected_cols if col not in actual_cols]
#             if missing_cols:
#                 error_txt = f"\n{excel_sheet_name} sheet is missing one or columns:\n{missing_cols}"
#                 raise config.MissingHeaderError(error_txt)
#             extra_columns = [col for col in actual_cols if col not in expected_cols]
#             if extra_columns:
#                 click.echo(f'\nWARNING: unneeded columns found on {excel_sheet_name} sheet:\n{extra_columns}')
#
#         # --- Import excel info dataframe -------------------------------------
#         self.df = pd.read_excel(
#             io=excel_path,
#             sheet_name=excel_sheet_name,
#             header=header_row-1,
#             converters=converters,
#             usecols=converters.keys() if converters else None,
#         )
#         self.path = excel_path
#         self.sheetname = excel_sheet_name
#         self.header_row = header_row
#         self.group = excel_sheet_name if group is None else group
#         self.year = year
#
#         if autosetup:
#             self.add_row_column()
#             self.drop_dups()
#             self.error_check()
#             self.reset_index()
#

#
#     def reset_index(self):
#         self.df = self.df.reset_index(drop=True)
#
#     def error_check(self):
#         # TODO - implement
#         #   raise MMError for "deal breakers"
#         #   Show warnings for all other "errors"
#         pass


if __name__ == '__main__':
    pass
