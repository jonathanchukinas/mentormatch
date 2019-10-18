# --- Standard Library Imports ------------------------------------------------
from pathlib import Path
from collections import Counter

# --- Third Party Imports -----------------------------------------------------
import openpyxl

# --- Intra-Package Imports ---------------------------------------------------
from mentormatch.main.exceptions import MentormatchError
from mentormatch.worksheet.get_path import get_path
from mentormatch.worksheet.schema import schemas


# main callable
def get_applications():

    groups = 'mentors mentees'.split()
    path = get_path()
    wb = get_workbook(path)
    applications = dict()
    for group in groups:
        ws = get_worksheet(wb, group)
        application_count = ws.max_row
        required_fields = schemas(group)
        check_for_missing_and_duplicate_fields(
            group=group,
            required_field_names=[field.name for field in required_fields],
            actual_field_names=get_field_names(ws)
        )
        # validated_fields = dict()
        # for field in required_fields:
        #     name = field.name
        #     validated_fields[name] = get_validated_values(ws, name, field.val_func)
        validated_fields = {
            field.name: get_validated_values(ws, field.name, field.val_func)
            for field
            in required_fields
        }
        applications[group] = validated_fields
    return applications


def check_for_missing_and_duplicate_fields(group, required_field_names, actual_field_names):
    actual_counter = Counter(actual_field_names)
    # required_count = dict()
    # for field in required_field_names:
    #     required_count[field] = actual_counter[field]
    required_count = {field: actual_counter[field] for field in required_field_names}
    required_missing = [field for field in required_field_names if required_count[field] == 0]
    required_duplicate = [field for field in required_field_names if required_count[field] > 1]
    msg = []
    if required_missing:
        msg.append(f'The following fields are missing from the {group} worksheet: {required_missing}')
    if required_duplicate:
        msg.append(f'The following fields have duplicated in the {group} worksheet: {required_duplicate}')
    if msg:
        raise MentormatchError('\n'.join(msg))


# Test complete
def get_workbook(path: Path):
    # TODO introduce error handling

    # --- Check that a file was selected --------------------------------------
    if str(path) == '.':
        raise MentormatchError("You didn't select a file")

    # --- Check that file has proper extension --------------------------------
    required_extensions = '.xlsx .xls'.split()
    if path.suffix not in required_extensions:
        msg = f"You selected a file without aproper extension: {required_extensions}"
        raise MentormatchError(msg)

    return openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)


# Test complete
def get_worksheet(workbook, sheetname):
    worksheets = dict()
    try:
        return workbook[sheetname]
    except KeyError:
        raise MentormatchError(f'Ensure excel workbook contains worksheet {sheetname}')


# Test complete
def get_field_names(worksheet):
    ws = worksheet
    return [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]


def get_validated_values(worksheet, field_name, validation_function):
    values = []
    col = get_field_column_number(worksheet, field_name)
    for row in range(2, worksheet.max_row+1):
        orig_value = worksheet.cell(row, col).value
        validated_value = validation_function(orig_value)
        values.append(validated_value)
    return values


# Test complete
def get_field_column_number(worksheet, field_name):
    try:
        return get_field_names(worksheet).index(field_name)
    except ValueError:
        raise ValueError(f'{field_name} not found in {worksheet}')


if __name__ == '__main__':
    mylist = '1 2 3'.split()
    pos = mylist.index(2)
    print(pos)
